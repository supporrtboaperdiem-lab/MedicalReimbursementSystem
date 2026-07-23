import html
import re
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from statistics import mean
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    from rapidfuzz import fuzz, process

    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    from difflib import SequenceMatcher

    RAPIDFUZZ_AVAILABLE = False


EXCEL_IMPORTER_VERSION = "2026-07-22-v11-hybrid-intelligent"


class ExcelPriceListService:
    """
    Hybrid Excel price-list importer.

    Detection strategy
    ------------------
    1. Exact normalized header matching.
    2. Safe fuzzy header matching for spelling and wording variations.
    3. Statistical profiling of the data below each potential header.
    4. Confidence scoring for service, unit, category, code, and price columns.
    5. Price relationship analysis when multiple numeric price columns exist.
    6. Strict full-cell numeric parsing to prevent service names such as
       "COVID-19", "CA-15.3", or "Adult Diaper 30 Pcs" becoming prices.

    Price policy
    ------------
    The preferred payable price is selected in this order:

        1. Approved / negotiated / contract / agreed price
        2. Discounted / net / final / corporate / member price
        3. Exact unit price
        4. Normal price / amount / rate / cost / tariff / charge

    If a preferred price cell is empty or invalid on a particular row, the
    importer falls back to the next valid recognized price column for that row.
    """

    MAX_HEADER_ROWS = 120
    MAX_COLUMNS = 120
    MAX_DATA_ROWS_PER_SHEET = 100_000
    EMPTY_ROW_STOP_LIMIT = 150

    PROFILE_SAMPLE_ROWS = 150
    MIN_PROFILE_POPULATED = 3
    MIN_HEADER_SCHEMA_SCORE = 45.0
    MIN_INFERRED_SCHEMA_SCORE = 62.0
    MIN_PRICE_NUMERIC_RATIO = 0.55
    MIN_SERVICE_TEXT_RATIO = 0.55
    FUZZY_HEADER_THRESHOLD = 84.0

    SERVICE_HEADERS = {
        "name",
        "service",
        "services",
        "service name",
        "service names",
        "service description",
        "medical service",
        "medical services",
        "item",
        "items",
        "item name",
        "item names",
        "item description",
        "description",
        "descriptions",
        "particular",
        "particulars",
        "procedure",
        "procedures",
        "procedure name",
        "procedure names",
        "procedure description",
        "procedure descriptions",
        "activity",
        "activities",
        "activity name",
        "activity description",
        "investigation",
        "investigations",
        "investigation name",
        "investigation description",
        "treatment",
        "treatments",
        "test",
        "tests",
        "test name",
        "test names",
        "test description",
        "drug name",
        "medicine name",
        "product name",
        "charge item",
        "charge description",
        "benefit",
        "benefit description",
        "types of main activities",
        "types of main activity",
        "types of main activities procedures",
        "types of main activity procedures",
        "types of activities procedures",
        "laboratory service",
        "laboratory services",
        "laboratory test",
        "laboratory tests",
        "lab service",
        "lab services",
        "lab test",
        "lab tests",
        "types of main laboratory test",
        "types of main laboratory tests",
        "types of laboratory test",
        "types of laboratory tests",
        "main laboratory test",
        "main laboratory tests",
        "x ray",
        "x rays",
        "x ray service",
        "x ray services",
        "x ray examination",
        "x ray examinations",
        "x ray procedure",
        "x ray procedures",
        "diagnostic radiology",
        "diagnostic radiology service",
        "diagnostic radiology services",
        "diagnostic radiology examination",
        "diagnostic radiology examinations",
        "types of main diagnostic radiology",
        "types of diagnostic radiology",
        "main diagnostic radiology",
        "diognostic radiology",
        "diognostic radiology service",
        "diognostic radiology services",
        "diognostic radiology examination",
        "diognostic radiology examinations",
        "types of main diognostic radiology",
        "types of diognostic radiology",
        "main diognostic radiology",
        "radiology",
        "radiology service",
        "radiology services",
        "radiology examination",
        "radiology examinations",
        "radiology procedure",
        "radiology procedures",
        "imaging",
        "imaging service",
        "imaging services",
        "imaging examination",
        "imaging examinations",
        "imaging procedure",
        "imaging procedures",
    }

    UNIT_HEADERS = {
        "unit",
        "units",
        "uom",
        "u m",
        "unit of measure",
        "unit of measurement",
        "measurement unit",
        "measure",
        "package",
        "pack",
        "pack size",
        "dosage form",
        "form",
        "service unit",
        "billing unit",
    }

    CATEGORY_HEADERS = {
        "category",
        "categories",
        "service category",
        "item category",
        "type",
        "service type",
        "item type",
        "department",
        "section",
        "group",
        "class",
        "classification",
        "subcategory",
        "sub category",
    }

    CODE_HEADERS = {
        "code",
        "service code",
        "item code",
        "procedure code",
        "test code",
        "product code",
        "billing code",
        "tariff code",
        "reference code",
    }

    SERIAL_HEADERS = {
        "s n",
        "sn",
        "s no",
        "serial",
        "serial no",
        "serial number",
        "no",
        "number",
        "row",
        "row no",
    }

    APPROVED_PRICE_HEADERS = {
        "approved price",
        "negotiated price",
        "contract price",
        "agreed price",
        "boa price",
        "bank of abyssinia price",
        "bank price",
        "insurance approved price",
        "payable price",
        "agreed tariff",
        "contract tariff",
        "negotiated tariff",
    }

    DISCOUNT_PRICE_HEADERS = {
        "discounted price",
        "discount price",
        "final price",
        "net price",
        "selling price after discount",
        "price after discount",
        "corporate price",
        "corporate rate",
        "member price",
        "member rate",
        "scheme price",
        "scheme rate",
        "partner price",
        "partner rate",
        "insurance price",
        "insurance rate",
        "special price",
    }

    EXACT_UNIT_PRICE_HEADERS = {
        "unit price",
        "unit price etb",
        "unit price birr",
        "unit price br",
        "price per unit",
        "unit cost",
        "unit rate",
    }

    NORMAL_PRICE_HEADERS = {
        "price",
        "price etb",
        "price birr",
        "price br",
        "regular price",
        "cash price",
        "selling price",
        "list price",
        "gross price",
        "amount",
        "rate",
        "cost",
        "tariff",
        "charge",
        "fee",
        "hospital price",
        "hospital rate",
        "standard price",
    }

    IGNORE_SERVICE_VALUES = {
        "",
        "none",
        "n a",
        "na",
        "n/a",
        "-",
        "--",
        "service",
        "services",
        "service name",
        "service description",
        "medical service",
        "item",
        "item name",
        "item description",
        "description",
        "procedure",
        "procedure name",
        "activity",
        "investigation",
        "treatment",
        "category",
        "subcategory",
        "code",
        "no",
        "number",
        "unit",
        "units",
        "uom",
        "price",
        "unit price",
        "discounted price",
        "discount price",
        "approved price",
        "total",
        "subtotal",
        "grand total",
    }

    IGNORE_UNIT_VALUES = {
        "",
        "none",
        "unit",
        "units",
        "uom",
        "n a",
        "na",
        "n/a",
        "-",
        "--",
    }

    PRICE_ROLE_PRIORITY = {
        "approved_price": 0,
        "discounted_price": 1,
        "unit_price": 2,
        "price": 3,
        "inferred_price": 4,
    }

    HEADER_ALIASES = {
        "service": SERVICE_HEADERS,
        "unit": UNIT_HEADERS,
        "category": CATEGORY_HEADERS,
        "code": CODE_HEADERS,
        "serial": SERIAL_HEADERS,
        "approved_price": APPROVED_PRICE_HEADERS,
        "discounted_price": DISCOUNT_PRICE_HEADERS,
        "unit_price": EXACT_UNIT_PRICE_HEADERS,
        "price": NORMAL_PRICE_HEADERS,
    }

    @staticmethod
    def normalize_header(value: Any) -> str:
        if value is None:
            return ""

        text = html.unescape(str(value))
        text = (
            text.replace("\n", " ")
            .replace("\r", " ")
            .replace("\t", " ")
            .replace("_", " ")
            .replace("\u00a0", " ")
            .replace("\u202f", " ")
        )
        text = re.sub(r"[^a-zA-Z0-9%]+", " ", text)
        return re.sub(r"\s+", " ", text).strip().lower()

    @staticmethod
    def normalize_text(value: Any) -> str:
        if value is None:
            return ""

        text = html.unescape(str(value))
        text = (
            text.replace("\n", " ")
            .replace("\r", " ")
            .replace("\t", " ")
            .replace("\u00a0", " ")
            .replace("\u202f", " ")
        )
        return re.sub(r"\s+", " ", text).strip()

    @staticmethod
    def to_decimal(value: Any) -> Optional[Decimal]:
        """
        Parse only a fully numeric or currency-formatted cell.

        Accepted:
            850
            850.00
            "1,200.00"
            "ETB 1,200.00"
            "1,200.00 Birr"
            "(500.00)"

        Rejected:
            "Covid-19"
            "CA-15.3"
            "2nd Opinion CT Reading"
            "Adult Diaper Large Size Of 30 Pcs"
        """
        if value is None or isinstance(value, bool):
            return None

        if isinstance(value, Decimal):
            return value

        if isinstance(value, int):
            return Decimal(value)

        if isinstance(value, float):
            try:
                return Decimal(str(value))
            except (InvalidOperation, ValueError):
                return None

        text = html.unescape(str(value)).strip()
        if not text or text.startswith("="):
            return None

        text = text.replace("\u00a0", " ").replace("\u202f", " ")

        numeric_pattern = re.compile(
            r"""
            ^\s*
            (?P<open>\()?
            \s*
            (?:(?:ETB|BIRR|BR|USD)\s*)?
            (?P<number>
                -?
                (?:
                    \d{1,3}(?:,\d{3})+
                    |
                    \d+
                )
                (?:\.\d+)?
            )
            \s*
            (?:(?:ETB|BIRR|BR|USD))?
            \s*
            (?P<close>\))?
            \s*$
            """,
            re.IGNORECASE | re.VERBOSE,
        )

        match = numeric_pattern.fullmatch(text)
        if not match:
            return None

        number = match.group("number").replace(",", "")

        try:
            result = Decimal(number)
        except InvalidOperation:
            return None

        if match.group("open") and match.group("close"):
            result = -abs(result)

        return result

    @staticmethod
    def _is_textual(value: Any) -> bool:
        if value is None:
            return False
        if isinstance(value, bool):
            return False
        if isinstance(value, (int, float, Decimal)):
            return False
        return bool(str(value).strip())

    @staticmethod
    def _looks_like_code(value: Any) -> bool:
        if value is None:
            return False

        text = str(value).strip()
        if not text or len(text) > 40 or " " in text:
            return False

        if not re.fullmatch(r"[A-Za-z0-9._/-]+", text):
            return False

        has_letter = bool(re.search(r"[A-Za-z]", text))
        has_digit = bool(re.search(r"\d", text))
        return has_letter and has_digit

    @staticmethod
    def _looks_like_serial(value: Any) -> bool:
        if isinstance(value, bool) or value is None:
            return False

        if isinstance(value, int):
            return value >= 0

        if isinstance(value, float):
            return value >= 0 and value.is_integer()

        text = str(value).strip()
        return bool(re.fullmatch(r"\d+", text))

    @classmethod
    def _exact_header_role(cls, value: Any) -> Optional[str]:
        header = cls.normalize_header(value)
        if not header:
            return None

        for role, aliases in cls.HEADER_ALIASES.items():
            if header in aliases:
                return role

        return None

    @staticmethod
    def _fuzzy_ratio(left: str, right: str) -> float:
        if RAPIDFUZZ_AVAILABLE:
            return float(fuzz.token_set_ratio(left, right))

        return SequenceMatcher(None, left, right).ratio() * 100.0

    @classmethod
    def _fuzzy_header_role(
        cls,
        value: Any,
    ) -> Optional[Dict[str, Any]]:
        header = cls.normalize_header(value)
        if not header:
            return None

        # Long data values are not credible headers.
        if len(header) > 70 or len(header.split()) > 9:
            return None

        best: Optional[Dict[str, Any]] = None

        for role, aliases in cls.HEADER_ALIASES.items():
            if RAPIDFUZZ_AVAILABLE:
                result = process.extractOne(
                    header,
                    aliases,
                    scorer=fuzz.token_set_ratio,
                )
                if result is None:
                    continue
                alias, score, _ = result
            else:
                alias = ""
                score = 0.0
                for candidate in aliases:
                    candidate_score = cls._fuzzy_ratio(
                        header,
                        candidate,
                    )
                    if candidate_score > score:
                        score = candidate_score
                        alias = candidate

            if best is None or score > best["score"]:
                best = {
                    "role": role,
                    "alias": alias,
                    "score": float(score),
                }

        if best and best["score"] >= cls.FUZZY_HEADER_THRESHOLD:
            return best

        return None

    @classmethod
    def classify_header(
        cls,
        value: Any,
    ) -> Optional[Dict[str, Any]]:
        exact_role = cls._exact_header_role(value)

        if exact_role:
            return {
                "role": exact_role,
                "confidence": 1.0,
                "method": "exact",
                "matched_alias": cls.normalize_header(value),
            }

        fuzzy_match = cls._fuzzy_header_role(value)
        if fuzzy_match:
            return {
                "role": fuzzy_match["role"],
                "confidence": fuzzy_match["score"] / 100.0,
                "method": "fuzzy",
                "matched_alias": fuzzy_match["alias"],
            }

        return None

    @classmethod
    def _profile_column(
        cls,
        worksheet: Worksheet,
        column_number: int,
        start_row: int,
        max_rows: Optional[int] = None,
    ) -> Dict[str, Any]:
        sample_rows = max_rows or cls.PROFILE_SAMPLE_ROWS
        end_row = min(
            worksheet.max_row or start_row,
            start_row + sample_rows - 1,
        )

        populated_values: List[Any] = []
        numeric_values: List[Decimal] = []
        text_values: List[str] = []
        code_count = 0
        serial_count = 0
        positive_numeric_count = 0
        integer_numeric_count = 0

        for row_number in range(start_row, end_row + 1):
            value = worksheet.cell(
                row=row_number,
                column=column_number,
            ).value

            if value is None or str(value).strip() == "":
                continue

            populated_values.append(value)
            parsed = cls.to_decimal(value)

            if parsed is not None:
                numeric_values.append(parsed)

                if parsed > 0:
                    positive_numeric_count += 1

                if parsed == parsed.to_integral_value():
                    integer_numeric_count += 1
            else:
                text = cls.normalize_text(value)
                text_values.append(text)

                if cls._looks_like_code(value):
                    code_count += 1

                if cls._looks_like_serial(value):
                    serial_count += 1

        populated_count = len(populated_values)
        numeric_count = len(numeric_values)
        text_count = len(text_values)

        numeric_ratio = (
            numeric_count / populated_count
            if populated_count
            else 0.0
        )
        text_ratio = (
            text_count / populated_count
            if populated_count
            else 0.0
        )
        positive_ratio = (
            positive_numeric_count / numeric_count
            if numeric_count
            else 0.0
        )
        integer_ratio = (
            integer_numeric_count / numeric_count
            if numeric_count
            else 0.0
        )
        unique_ratio = (
            len({str(value).strip().casefold() for value in populated_values})
            / populated_count
            if populated_count
            else 0.0
        )
        average_text_length = (
            mean(len(value) for value in text_values)
            if text_values
            else 0.0
        )
        code_ratio = (
            code_count / populated_count
            if populated_count
            else 0.0
        )
        serial_ratio = (
            serial_count / populated_count
            if populated_count
            else 0.0
        )

        return {
            "column": column_number,
            "start_row": start_row,
            "end_row": end_row,
            "populated_count": populated_count,
            "numeric_count": numeric_count,
            "text_count": text_count,
            "numeric_ratio": numeric_ratio,
            "text_ratio": text_ratio,
            "positive_ratio": positive_ratio,
            "integer_ratio": integer_ratio,
            "unique_ratio": unique_ratio,
            "average_text_length": average_text_length,
            "code_ratio": code_ratio,
            "serial_ratio": serial_ratio,
            "sample_values": [
                cls.normalize_text(value)
                for value in populated_values[:8]
            ],
            "numeric_values": numeric_values,
        }

    @classmethod
    def _service_profile_score(
        cls,
        profile: Dict[str, Any],
    ) -> float:
        if profile["populated_count"] < cls.MIN_PROFILE_POPULATED:
            return 0.0

        score = 0.0
        score += profile["text_ratio"] * 45.0
        score += profile["unique_ratio"] * 25.0

        average_length = profile["average_text_length"]
        if 5 <= average_length <= 120:
            score += 18.0
        elif average_length > 2:
            score += 8.0

        score -= profile["code_ratio"] * 20.0
        score -= profile["serial_ratio"] * 25.0
        score -= profile["numeric_ratio"] * 35.0
        return max(0.0, score)

    @classmethod
    def _category_profile_score(
        cls,
        profile: Dict[str, Any],
    ) -> float:
        if profile["populated_count"] < cls.MIN_PROFILE_POPULATED:
            return 0.0

        score = profile["text_ratio"] * 35.0

        if 0.02 <= profile["unique_ratio"] <= 0.40:
            score += 35.0

        if 2 <= profile["average_text_length"] <= 50:
            score += 15.0

        score -= profile["numeric_ratio"] * 35.0
        score -= profile["code_ratio"] * 15.0
        return max(0.0, score)

    @classmethod
    def _code_profile_score(
        cls,
        profile: Dict[str, Any],
    ) -> float:
        if profile["populated_count"] < cls.MIN_PROFILE_POPULATED:
            return 0.0

        score = profile["code_ratio"] * 70.0
        score += profile["unique_ratio"] * 20.0
        score -= profile["numeric_ratio"] * 20.0
        return max(0.0, score)

    @classmethod
    def _serial_profile_score(
        cls,
        profile: Dict[str, Any],
    ) -> float:
        if profile["populated_count"] < cls.MIN_PROFILE_POPULATED:
            return 0.0

        score = profile["serial_ratio"] * 65.0
        score += profile["integer_ratio"] * 20.0
        score -= profile["text_ratio"] * 30.0
        return max(0.0, score)

    @classmethod
    def _price_profile_score(
        cls,
        profile: Dict[str, Any],
    ) -> float:
        if profile["populated_count"] < cls.MIN_PROFILE_POPULATED:
            return 0.0

        score = profile["numeric_ratio"] * 55.0
        score += profile["positive_ratio"] * 25.0
        score += profile["unique_ratio"] * 8.0

        # Pure serial-number columns are poor price candidates.
        score -= profile["serial_ratio"] * 35.0

        if profile["numeric_values"]:
            median_like = sorted(
                profile["numeric_values"]
            )[len(profile["numeric_values"]) // 2]

            if median_like > 1:
                score += 8.0

        return max(0.0, score)

    @classmethod
    def _unit_profile_score(
        cls,
        profile: Dict[str, Any],
    ) -> float:
        if profile["populated_count"] < cls.MIN_PROFILE_POPULATED:
            return 0.0

        score = profile["text_ratio"] * 25.0

        if 0.01 <= profile["unique_ratio"] <= 0.55:
            score += 20.0

        if 1 <= profile["average_text_length"] <= 30:
            score += 15.0

        score -= profile["numeric_ratio"] * 25.0
        return max(0.0, score)

    @classmethod
    def _price_relationship_score(
        cls,
        candidate_values: Sequence[Decimal],
        reference_values: Sequence[Decimal],
    ) -> float:
        comparable = min(
            len(candidate_values),
            len(reference_values),
        )
        if comparable < 5:
            return 0.0

        lower_or_equal = 0
        ratio_matches = 0

        for candidate, reference in zip(
            candidate_values[:comparable],
            reference_values[:comparable],
        ):
            if candidate <= 0 or reference <= 0:
                continue

            if candidate <= reference:
                lower_or_equal += 1

            ratio = candidate / reference
            if Decimal("0.40") <= ratio <= Decimal("1.00"):
                ratio_matches += 1

        return (
            (lower_or_equal / comparable) * 12.0
            + (ratio_matches / comparable) * 8.0
        )

    @classmethod
    def _candidate_from_header(
        cls,
        worksheet: Worksheet,
        row_number: int,
    ) -> Optional[Dict[str, Any]]:
        max_column = min(
            worksheet.max_column or 0,
            cls.MAX_COLUMNS,
        )

        recognized: Dict[
            str,
            List[Dict[str, Any]],
        ] = {}

        non_empty_count = 0

        for column_number in range(1, max_column + 1):
            raw_value = worksheet.cell(
                row=row_number,
                column=column_number,
            ).value

            if cls.normalize_text(raw_value):
                non_empty_count += 1

            classification = cls.classify_header(raw_value)
            if classification is None:
                continue

            role = classification["role"]

            recognized.setdefault(role, []).append(
                {
                    "column": column_number,
                    "header": cls.normalize_header(raw_value),
                    "raw_header": raw_value,
                    "header_confidence": classification["confidence"],
                    "header_method": classification["method"],
                    "matched_alias": classification["matched_alias"],
                    "role": role,
                }
            )

        service_candidates = recognized.get("service", [])
        raw_price_candidates: List[Dict[str, Any]] = []

        for role in (
            "approved_price",
            "discounted_price",
            "unit_price",
            "price",
        ):
            for candidate in recognized.get(role, []):
                raw_price_candidates.append(
                    {
                        **candidate,
                        "priority": cls.PRICE_ROLE_PRIORITY[role],
                    }
                )

        if not service_candidates or not raw_price_candidates:
            return None

        profiles = {
            column_number: cls._profile_column(
                worksheet,
                column_number,
                row_number + 1,
            )
            for column_number in range(1, max_column + 1)
        }

        for candidate in service_candidates:
            profile = profiles[candidate["column"]]
            candidate["profile"] = profile
            candidate["profile_score"] = (
                cls._service_profile_score(profile)
            )
            candidate["total_score"] = (
                candidate["header_confidence"] * 45.0
                + candidate["profile_score"]
            )

        price_candidates: List[Dict[str, Any]] = []

        for candidate in raw_price_candidates:
            profile = profiles[candidate["column"]]
            candidate["profile"] = profile
            candidate["profile_score"] = (
                cls._price_profile_score(profile)
            )
            candidate["total_score"] = (
                candidate["header_confidence"] * 40.0
                + candidate["profile_score"]
                + max(
                    0.0,
                    12.0 - candidate["priority"] * 3.0,
                )
            )

            if (
                profile["numeric_ratio"]
                >= cls.MIN_PRICE_NUMERIC_RATIO
            ):
                price_candidates.append(candidate)

        if not price_candidates:
            return None

        service_candidates.sort(
            key=lambda item: (
                item["total_score"],
                -item["column"],
            ),
            reverse=True,
        )

        price_candidates.sort(
            key=lambda item: (
                item["priority"],
                -item["total_score"],
                item["column"],
            )
        )

        selected_service = service_candidates[0]

        if (
            selected_service["profile"]["text_ratio"]
            < cls.MIN_SERVICE_TEXT_RATIO
        ):
            return None

        category_candidate = cls._select_optional_role(
            recognized.get("category", []),
            profiles,
            cls._category_profile_score,
        )
        code_candidate = cls._select_optional_role(
            recognized.get("code", []),
            profiles,
            cls._code_profile_score,
        )
        unit_candidate = cls._select_optional_role(
            recognized.get("unit", []),
            profiles,
            cls._unit_profile_score,
        )
        serial_candidate = cls._select_optional_role(
            recognized.get("serial", []),
            profiles,
            cls._serial_profile_score,
        )

        schema_score = (
            selected_service["total_score"]
            + price_candidates[0]["total_score"]
            + min(non_empty_count, 8) * 2.0
        )

        if code_candidate:
            schema_score += 5.0
        if category_candidate:
            schema_score += 5.0
        if unit_candidate:
            schema_score += 3.0
        if serial_candidate:
            schema_score += 2.0

        return {
            "method": "header",
            "schema_score": schema_score,
            "header_row": row_number,
            "service_column": selected_service["column"],
            "service_header": selected_service["header"],
            "service_confidence": min(
                1.0,
                selected_service["total_score"] / 100.0,
            ),
            "unit_column": (
                unit_candidate["column"]
                if unit_candidate
                else None
            ),
            "unit_header": (
                unit_candidate["header"]
                if unit_candidate
                else None
            ),
            "category_column": (
                category_candidate["column"]
                if category_candidate
                else None
            ),
            "category_header": (
                category_candidate["header"]
                if category_candidate
                else None
            ),
            "code_column": (
                code_candidate["column"]
                if code_candidate
                else None
            ),
            "code_header": (
                code_candidate["header"]
                if code_candidate
                else None
            ),
            "serial_column": (
                serial_candidate["column"]
                if serial_candidate
                else None
            ),
            "price_candidates": price_candidates,
            "price_columns": [
                candidate["column"]
                for candidate in price_candidates
            ],
            "profiles": profiles,
        }

    @classmethod
    def _select_optional_role(
        cls,
        candidates: List[Dict[str, Any]],
        profiles: Dict[int, Dict[str, Any]],
        profile_scorer,
    ) -> Optional[Dict[str, Any]]:
        if not candidates:
            return None

        scored: List[Dict[str, Any]] = []

        for candidate in candidates:
            profile = profiles[candidate["column"]]
            total_score = (
                candidate["header_confidence"] * 35.0
                + profile_scorer(profile)
            )

            scored.append(
                {
                    **candidate,
                    "profile": profile,
                    "total_score": total_score,
                }
            )

        scored.sort(
            key=lambda item: item["total_score"],
            reverse=True,
        )
        return scored[0]

    @classmethod
    def _candidate_from_profiles(
        cls,
        worksheet: Worksheet,
        assumed_header_row: int,
    ) -> Optional[Dict[str, Any]]:
        """
        Last-resort inference for sheets with unfamiliar or missing headers.
        """
        max_column = min(
            worksheet.max_column or 0,
            cls.MAX_COLUMNS,
        )
        start_row = assumed_header_row + 1

        profiles = {
            column_number: cls._profile_column(
                worksheet,
                column_number,
                start_row,
            )
            for column_number in range(1, max_column + 1)
        }

        service_ranked = sorted(
            (
                {
                    "column": column_number,
                    "score": cls._service_profile_score(profile),
                    "profile": profile,
                }
                for column_number, profile in profiles.items()
            ),
            key=lambda item: item["score"],
            reverse=True,
        )

        price_ranked = sorted(
            (
                {
                    "column": column_number,
                    "score": cls._price_profile_score(profile),
                    "profile": profile,
                }
                for column_number, profile in profiles.items()
            ),
            key=lambda item: item["score"],
            reverse=True,
        )

        if not service_ranked or not price_ranked:
            return None

        selected_service = service_ranked[0]

        price_ranked = [
            candidate
            for candidate in price_ranked
            if candidate["column"] != selected_service["column"]
            and candidate["profile"]["numeric_ratio"]
            >= cls.MIN_PRICE_NUMERIC_RATIO
        ]

        if not price_ranked:
            return None

        price_candidates: List[Dict[str, Any]] = []

        for index, candidate in enumerate(price_ranked[:4]):
            price_candidates.append(
                {
                    "column": candidate["column"],
                    "header": cls.normalize_header(
                        worksheet.cell(
                            row=assumed_header_row,
                            column=candidate["column"],
                        ).value
                    ),
                    "raw_header": worksheet.cell(
                        row=assumed_header_row,
                        column=candidate["column"],
                    ).value,
                    "header_confidence": 0.0,
                    "header_method": "profile",
                    "matched_alias": None,
                    "role": "inferred_price",
                    "priority": (
                        cls.PRICE_ROLE_PRIORITY["inferred_price"]
                        + index
                    ),
                    "profile": candidate["profile"],
                    "profile_score": candidate["score"],
                    "total_score": candidate["score"],
                }
            )

        schema_score = (
            selected_service["score"]
            + price_candidates[0]["total_score"]
        )

        if (
            selected_service["profile"]["text_ratio"]
            < cls.MIN_SERVICE_TEXT_RATIO
            or schema_score < cls.MIN_INFERRED_SCHEMA_SCORE
        ):
            return None

        excluded_columns = {
            selected_service["column"],
            *[
                candidate["column"]
                for candidate in price_candidates
            ],
        }

        optional_profiles = [
            (
                column_number,
                profile,
            )
            for column_number, profile in profiles.items()
            if column_number not in excluded_columns
        ]

        category_column = cls._best_profile_column(
            optional_profiles,
            cls._category_profile_score,
            minimum_score=40.0,
        )
        code_column = cls._best_profile_column(
            optional_profiles,
            cls._code_profile_score,
            minimum_score=45.0,
        )
        unit_column = cls._best_profile_column(
            optional_profiles,
            cls._unit_profile_score,
            minimum_score=40.0,
        )

        return {
            "method": "profile",
            "schema_score": schema_score,
            "header_row": assumed_header_row,
            "service_column": selected_service["column"],
            "service_header": cls.normalize_header(
                worksheet.cell(
                    row=assumed_header_row,
                    column=selected_service["column"],
                ).value
            ),
            "service_confidence": min(
                1.0,
                selected_service["score"] / 100.0,
            ),
            "unit_column": unit_column,
            "unit_header": cls.normalize_header(
                worksheet.cell(
                    row=assumed_header_row,
                    column=unit_column,
                ).value
            )
            if unit_column
            else None,
            "category_column": category_column,
            "category_header": cls.normalize_header(
                worksheet.cell(
                    row=assumed_header_row,
                    column=category_column,
                ).value
            )
            if category_column
            else None,
            "code_column": code_column,
            "code_header": cls.normalize_header(
                worksheet.cell(
                    row=assumed_header_row,
                    column=code_column,
                ).value
            )
            if code_column
            else None,
            "serial_column": None,
            "price_candidates": price_candidates,
            "price_columns": [
                candidate["column"]
                for candidate in price_candidates
            ],
            "profiles": profiles,
        }

    @staticmethod
    def _best_profile_column(
        profile_items: Iterable[
            Tuple[int, Dict[str, Any]]
        ],
        scorer,
        minimum_score: float,
    ) -> Optional[int]:
        scored = [
            (column_number, scorer(profile))
            for column_number, profile in profile_items
        ]

        if not scored:
            return None

        column_number, score = max(
            scored,
            key=lambda item: item[1],
        )

        return (
            column_number
            if score >= minimum_score
            else None
        )

    @classmethod
    def _reorder_inferred_price_candidates(
        cls,
        price_candidates: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """
        When headers are unknown, prefer a consistently lower numeric column
        over a higher one because it is often the negotiated/discounted price.
        """
        if len(price_candidates) < 2:
            return price_candidates

        rescored = []

        for candidate in price_candidates:
            score = candidate["total_score"]

            for reference in price_candidates:
                if candidate["column"] == reference["column"]:
                    continue

                score += cls._price_relationship_score(
                    candidate["profile"]["numeric_values"],
                    reference["profile"]["numeric_values"],
                )

            rescored.append(
                {
                    **candidate,
                    "relationship_score": score,
                }
            )

        rescored.sort(
            key=lambda item: (
                -item["relationship_score"],
                item["column"],
            )
        )

        for index, candidate in enumerate(rescored):
            candidate["priority"] = (
                cls.PRICE_ROLE_PRIORITY["inferred_price"]
                + index
            )

        return rescored

    @classmethod
    def detect_table(
        cls,
        worksheet: Worksheet,
    ) -> Optional[Dict[str, Any]]:
        print(
            f"[EXCEL] Detecting table in sheet: "
            f"{worksheet.title}",
            flush=True,
        )

        max_row = min(
            worksheet.max_row or 0,
            cls.MAX_HEADER_ROWS,
        )
        max_column = min(
            worksheet.max_column or 0,
            cls.MAX_COLUMNS,
        )

        if max_row <= 0 or max_column <= 0:
            return None

        candidates: List[Dict[str, Any]] = []

        # First pass: recognized headers.
        for row_number in range(1, max_row + 1):
            candidate = cls._candidate_from_header(
                worksheet,
                row_number,
            )

            if candidate is not None:
                candidates.append(candidate)

        if candidates:
            candidates.sort(
                key=lambda item: (
                    item["schema_score"],
                    -item["header_row"],
                ),
                reverse=True,
            )
            best = candidates[0]

            if best["schema_score"] >= cls.MIN_HEADER_SCHEMA_SCORE:
                best["price_candidates"].sort(
                    key=lambda item: (
                        item["priority"],
                        -item["total_score"],
                        item["column"],
                    )
                )

                print(
                    f"[EXCEL][SCHEMA] "
                    f"Sheet={worksheet.title!r}, "
                    f"Method={best['method']!r}, "
                    f"Score={best['schema_score']:.2f}, "
                    f"HeaderRow={best['header_row']}, "
                    f"ServiceColumn={best['service_column']}, "
                    f"UnitColumn={best['unit_column']}, "
                    f"CategoryColumn="
                    f"{best['category_column']}, "
                    f"CodeColumn={best['code_column']}, "
                    f"PriceCandidates="
                    f"{cls._serializable_price_candidates(best['price_candidates'])}",
                    flush=True,
                )

                return best

        # Second pass: data-profile inference.
        # Try likely header rows and choose the strongest inferred schema.
        inferred_candidates: List[Dict[str, Any]] = []

        for assumed_header_row in range(
            1,
            min(max_row, 30) + 1,
        ):
            inferred = cls._candidate_from_profiles(
                worksheet,
                assumed_header_row,
            )

            if inferred is not None:
                inferred["price_candidates"] = (
                    cls._reorder_inferred_price_candidates(
                        inferred["price_candidates"]
                    )
                )
                inferred_candidates.append(inferred)

        if inferred_candidates:
            inferred_candidates.sort(
                key=lambda item: (
                    item["schema_score"],
                    -item["header_row"],
                ),
                reverse=True,
            )
            best = inferred_candidates[0]

            print(
                f"[EXCEL][SCHEMA] "
                f"Sheet={worksheet.title!r}, "
                f"Method={best['method']!r}, "
                f"Score={best['schema_score']:.2f}, "
                f"HeaderRow={best['header_row']}, "
                f"ServiceColumn={best['service_column']}, "
                f"UnitColumn={best['unit_column']}, "
                f"CategoryColumn="
                f"{best['category_column']}, "
                f"CodeColumn={best['code_column']}, "
                f"PriceCandidates="
                f"{cls._serializable_price_candidates(best['price_candidates'])}",
                flush=True,
            )

            return best

        return None

    @staticmethod
    def _serializable_price_candidates(
        candidates: Sequence[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        return [
            {
                "column": candidate["column"],
                "header": candidate.get("header"),
                "role": candidate.get("role"),
                "priority": candidate.get("priority"),
                "header_method": candidate.get(
                    "header_method"
                ),
                "numeric_ratio": round(
                    candidate.get(
                        "profile",
                        {},
                    ).get(
                        "numeric_ratio",
                        0.0,
                    ),
                    3,
                ),
                "score": round(
                    candidate.get(
                        "total_score",
                        0.0,
                    ),
                    2,
                ),
            }
            for candidate in candidates
        ]

    @classmethod
    def build_service_name(
        cls,
        service_name: Any,
        unit_value: Any,
    ) -> str:
        service = cls.normalize_text(service_name)
        unit = cls.normalize_text(unit_value)

        if not service:
            return ""

        normalized_unit = cls.normalize_header(unit)

        if (
            not unit
            or normalized_unit in cls.IGNORE_UNIT_VALUES
        ):
            return service

        if unit.casefold() in service.casefold():
            return service

        return f"{service} ({unit})"

    @classmethod
    def _is_footer_or_note(
        cls,
        service_text: str,
    ) -> bool:
        normalized = cls.normalize_header(service_text)

        note_phrases = (
            "subject to service charge",
            "subject to 5 service charge",
            "subject to 5% service charge",
            "all prices are subject",
            "all price are subject",
            "grand total",
            "subtotal",
            "prepared by",
            "approved by",
            "effective date",
            "note",
        )

        return any(
            phrase in normalized
            for phrase in note_phrases
        )

    @classmethod
    def _select_row_price(
        cls,
        worksheet: Worksheet,
        row_number: int,
        price_candidates: Sequence[
            Dict[str, Any]
        ],
    ) -> Dict[str, Any]:
        attempts = []

        for candidate in price_candidates:
            cell = worksheet.cell(
                row=row_number,
                column=candidate["column"],
            )
            parsed = cls.to_decimal(cell.value)

            attempt = {
                "type": candidate.get("role"),
                "priority": candidate.get("priority"),
                "column": candidate["column"],
                "cell": cell.coordinate,
                "raw": cell.value,
                "parsed": parsed,
                "header": candidate.get("header"),
            }
            attempts.append(attempt)

            if parsed is not None and parsed > 0:
                return {
                    "approved_price": parsed,
                    "raw_price": cell.value,
                    "price_cell": cell.coordinate,
                    "price_column": candidate["column"],
                    "price_type": candidate.get("role"),
                    "price_header": candidate.get(
                        "header"
                    ),
                    "attempted": attempts,
                }

        return {
            "approved_price": None,
            "raw_price": None,
            "price_cell": None,
            "price_column": None,
            "price_type": None,
            "price_header": None,
            "attempted": attempts,
        }

    @classmethod
    def extract_from_worksheet(
        cls,
        worksheet: Worksheet,
        formula_worksheet: Optional[
            Worksheet
        ] = None,
    ) -> Tuple[
        List[Dict[str, Any]],
        Dict[str, Any],
    ]:
        table = cls.detect_table(worksheet)

        if table is None:
            print(
                f"[EXCEL] No reliable table found "
                f"in sheet: {worksheet.title}",
                flush=True,
            )

            return [], {
                "sheet": worksheet.title,
                "status": "SKIPPED",
                "error": (
                    "A reliable service and price "
                    "column combination was not found."
                ),
                "items_found": 0,
                "duplicates_skipped": 0,
                "invalid_rows_skipped": 0,
                "formula_price_rows": 0,
            }

        start_row = table["header_row"] + 1
        last_row = min(
            worksheet.max_row or start_row,
            start_row
            + cls.MAX_DATA_ROWS_PER_SHEET
            - 1,
        )

        items: List[Dict[str, Any]] = []
        seen_services: Dict[
            str,
            Dict[str, Any],
        ] = {}

        empty_rows = 0
        invalid_rows = 0
        duplicates = 0
        formula_price_rows = 0
        rows_examined = 0

        for row_number in range(
            start_row,
            last_row + 1,
        ):
            rows_examined += 1

            raw_service = worksheet.cell(
                row=row_number,
                column=table["service_column"],
            ).value
            service_text = cls.normalize_text(
                raw_service
            )

            raw_unit = (
                worksheet.cell(
                    row=row_number,
                    column=table["unit_column"],
                ).value
                if table["unit_column"]
                else None
            )
            unit_text = cls.normalize_text(raw_unit)

            raw_category = (
                worksheet.cell(
                    row=row_number,
                    column=table["category_column"],
                ).value
                if table["category_column"]
                else None
            )
            category_text = cls.normalize_text(
                raw_category
            )

            raw_code = (
                worksheet.cell(
                    row=row_number,
                    column=table["code_column"],
                ).value
                if table["code_column"]
                else None
            )
            code_text = cls.normalize_text(raw_code)

            price_result = cls._select_row_price(
                worksheet=worksheet,
                row_number=row_number,
                price_candidates=table[
                    "price_candidates"
                ],
            )

            all_price_cells_empty = all(
                attempt["raw"] is None
                or str(attempt["raw"]).strip() == ""
                for attempt in price_result["attempted"]
            )

            if (
                not service_text
                and not unit_text
                and not category_text
                and not code_text
                and all_price_cells_empty
            ):
                empty_rows += 1

                if empty_rows >= cls.EMPTY_ROW_STOP_LIMIT:
                    print(
                        f"[EXCEL] Stopped sheet "
                        f"{worksheet.title!r} after "
                        f"{cls.EMPTY_ROW_STOP_LIMIT} "
                        f"consecutive empty rows.",
                        flush=True,
                    )
                    break

                continue

            empty_rows = 0

            if not service_text:
                continue

            normalized_service = cls.normalize_header(
                service_text
            )

            if (
                normalized_service
                in cls.IGNORE_SERVICE_VALUES
            ):
                continue

            if cls._is_footer_or_note(service_text):
                continue

            # Repeated header rows inside the data.
            repeated_header = cls.classify_header(
                service_text
            )

            if (
                repeated_header is not None
                and repeated_header["role"] == "service"
            ):
                continue

            approved_price = price_result[
                "approved_price"
            ]

            if approved_price is None:
                formulas_found = []

                if formula_worksheet is not None:
                    for candidate in table[
                        "price_candidates"
                    ]:
                        formula_value = (
                            formula_worksheet.cell(
                                row=row_number,
                                column=candidate[
                                    "column"
                                ],
                            ).value
                        )

                        if (
                            isinstance(
                                formula_value,
                                str,
                            )
                            and formula_value.startswith(
                                "="
                            )
                        ):
                            formulas_found.append(
                                {
                                    "cell": (
                                        formula_worksheet
                                        .cell(
                                            row=row_number,
                                            column=candidate[
                                                "column"
                                            ],
                                        )
                                        .coordinate
                                    ),
                                    "formula": formula_value,
                                }
                            )

                if formulas_found:
                    formula_price_rows += 1

                    print(
                        f"[EXCEL][FORMULA PRICE WITHOUT "
                        f"CACHED RESULT] "
                        f"Sheet={worksheet.title!r}, "
                        f"Row={row_number}, "
                        f"Service={service_text!r}, "
                        f"Formulas={formulas_found}",
                        flush=True,
                    )
                else:
                    print(
                        f"[EXCEL][INVALID PRICE] "
                        f"Sheet={worksheet.title!r}, "
                        f"Row={row_number}, "
                        f"Service={service_text!r}, "
                        f"Candidates="
                        f"{price_result['attempted']}",
                        flush=True,
                    )

                invalid_rows += 1
                continue

            service_name = cls.build_service_name(
                service_text,
                unit_text,
            )

            if not service_name:
                invalid_rows += 1
                continue

            service_key = service_name.casefold()
            existing = seen_services.get(service_key)

            if existing is not None:
                duplicates += 1

                if (
                    existing["approved_price"]
                    != approved_price
                ):
                    existing.setdefault(
                        "conflicting_prices",
                        [],
                    ).append(
                        {
                            "price": approved_price,
                            "row": row_number,
                            "sheet": worksheet.title,
                            "price_cell": (
                                price_result[
                                    "price_cell"
                                ]
                            ),
                        }
                    )

                continue

            item = {
                "service_name": service_name,
                "original_service_name": (
                    service_text
                ),
                "code": code_text or None,
                "category": category_text or None,
                "unit": unit_text or None,
                "approved_price": approved_price,
                "price": approved_price,
                "source_sheet": worksheet.title,
                "source_row": row_number,
                "source_price_cell": (
                    price_result["price_cell"]
                ),
                "price_header": (
                    price_result["price_header"]
                ),
                "price_column": (
                    price_result["price_column"]
                ),
                "price_type": (
                    price_result["price_type"]
                ),
                "raw_price": (
                    price_result["raw_price"]
                ),
                "schema_method": table["method"],
                "schema_score": table[
                    "schema_score"
                ],
            }

            seen_services[service_key] = item
            items.append(item)

            print(
                f"[EXCEL][ACCEPTED] "
                f"Sheet={worksheet.title!r}, "
                f"Row={row_number}, "
                f"Service={service_name!r}, "
                f"Code={code_text!r}, "
                f"Category={category_text!r}, "
                f"PriceType="
                f"{price_result['price_type']!r}, "
                f"PriceCell="
                f"{price_result['price_cell']}, "
                f"RawPrice="
                f"{price_result['raw_price']!r}, "
                f"ParsedPrice={approved_price}",
                flush=True,
            )

        report = {
            "sheet": worksheet.title,
            "status": "PROCESSED",
            "schema_method": table["method"],
            "schema_score": table["schema_score"],
            "header_row": table["header_row"],
            "service_header": (
                table["service_header"]
            ),
            "service_column": (
                table["service_column"]
            ),
            "unit_header": table["unit_header"],
            "unit_column": table["unit_column"],
            "category_header": (
                table["category_header"]
            ),
            "category_column": (
                table["category_column"]
            ),
            "code_header": table["code_header"],
            "code_column": table["code_column"],
            "price_candidates": (
                cls._serializable_price_candidates(
                    table["price_candidates"]
                )
            ),
            "price_columns": table[
                "price_columns"
            ],
            "items_found": len(items),
            "duplicates_skipped": duplicates,
            "invalid_rows_skipped": invalid_rows,
            "formula_price_rows": (
                formula_price_rows
            ),
            "rows_examined": rows_examined,
        }

        print(
            f"[EXCEL] Sheet complete: "
            f"{worksheet.title}, "
            f"method={table['method']}, "
            f"score={table['schema_score']:.2f}, "
            f"items={len(items)}, "
            f"duplicates={duplicates}, "
            f"invalid={invalid_rows}, "
            f"formula_prices="
            f"{formula_price_rows}, "
            f"rows_examined={rows_examined}",
            flush=True,
        )

        return items, report

    @classmethod
    def extract_price_list(
        cls,
        file_path: Any,
    ) -> Dict[str, Any]:
        path = Path(file_path)

        if path.suffix.lower() != ".xlsx":
            raise ValueError(
                "Only XLSX Excel files are supported."
            )

        if not path.exists():
            raise FileNotFoundError(
                f"Excel file was not found: {path}"
            )

        print(
            f"[EXCEL] Importer version: "
            f"{EXCEL_IMPORTER_VERSION}",
            flush=True,
        )

        print(
            f"[EXCEL] RapidFuzz available: "
            f"{RAPIDFUZZ_AVAILABLE}",
            flush=True,
        )

        print(
            f"[EXCEL] Opening workbook: "
            f"{path.name}",
            flush=True,
        )

        workbook = load_workbook(
            filename=str(path),
            data_only=True,
            read_only=False,
            keep_links=False,
        )

        formula_workbook = load_workbook(
            filename=str(path),
            data_only=False,
            read_only=False,
            keep_links=False,
        )

        all_items: List[Dict[str, Any]] = []
        global_items: Dict[
            str,
            Dict[str, Any],
        ] = {}
        sheet_reports: List[
            Dict[str, Any]
        ] = []
        conflicts: List[
            Dict[str, Any]
        ] = []
        workbook_duplicates = 0

        try:
            worksheet_count = len(
                workbook.worksheets
            )

            print(
                f"[EXCEL] Workbook contains "
                f"{worksheet_count} worksheet(s).",
                flush=True,
            )

            for sheet_number, worksheet in enumerate(
                workbook.worksheets,
                start=1,
            ):
                print(
                    f"[EXCEL] Processing worksheet "
                    f"{sheet_number}/"
                    f"{worksheet_count}: "
                    f"{worksheet.title}",
                    flush=True,
                )

                formula_worksheet = (
                    formula_workbook[
                        worksheet.title
                    ]
                    if (
                        worksheet.title
                        in formula_workbook.sheetnames
                    )
                    else None
                )

                sheet_items, report = (
                    cls.extract_from_worksheet(
                        worksheet=worksheet,
                        formula_worksheet=(
                            formula_worksheet
                        ),
                    )
                )
                sheet_reports.append(report)

                for item in sheet_items:
                    service_key = (
                        item["service_name"]
                        .strip()
                        .casefold()
                    )
                    existing = global_items.get(
                        service_key
                    )

                    if existing is None:
                        global_items[
                            service_key
                        ] = item
                        all_items.append(item)
                        continue

                    workbook_duplicates += 1

                    if (
                        existing["approved_price"]
                        != item["approved_price"]
                    ):
                        conflict = {
                            "service_name": (
                                item["service_name"]
                            ),
                            "unit": item.get("unit"),
                            "first_price": (
                                existing[
                                    "approved_price"
                                ]
                            ),
                            "second_price": (
                                item["approved_price"]
                            ),
                            "first_sheet": (
                                existing[
                                    "source_sheet"
                                ]
                            ),
                            "second_sheet": (
                                item["source_sheet"]
                            ),
                            "first_row": (
                                existing["source_row"]
                            ),
                            "second_row": (
                                item["source_row"]
                            ),
                            "first_price_type": (
                                existing.get(
                                    "price_type"
                                )
                            ),
                            "second_price_type": (
                                item.get("price_type")
                            ),
                        }
                        conflicts.append(conflict)

                        print(
                            f"[EXCEL][WORKBOOK CONFLICT] "
                            f"{conflict}",
                            flush=True,
                        )

            processed_sheet_count = sum(
                1
                for report in sheet_reports
                if report.get("status")
                == "PROCESSED"
            )
            skipped_sheet_count = sum(
                1
                for report in sheet_reports
                if report.get("status")
                == "SKIPPED"
            )
            formula_price_row_count = sum(
                report.get(
                    "formula_price_rows",
                    0,
                )
                for report in sheet_reports
            )
            invalid_row_count = sum(
                report.get(
                    "invalid_rows_skipped",
                    0,
                )
                for report in sheet_reports
            )
            worksheet_item_total = sum(
                report.get(
                    "items_found",
                    0,
                )
                for report in sheet_reports
            )

            print(
                f"[EXCEL] Workbook extraction "
                f"completed. "
                f"Worksheets={worksheet_count}, "
                f"processed="
                f"{processed_sheet_count}, "
                f"skipped="
                f"{skipped_sheet_count}, "
                f"worksheet_items="
                f"{worksheet_item_total}, "
                f"unique_items="
                f"{len(all_items)}, "
                f"duplicates="
                f"{workbook_duplicates}, "
                f"conflicts="
                f"{len(conflicts)}, "
                f"invalid_rows="
                f"{invalid_row_count}, "
                f"formula_prices="
                f"{formula_price_row_count}",
                flush=True,
            )

            return {
                "items": all_items,
                "conflicts": conflicts,
                "sheets": sheet_reports,
                "sheet_reports": sheet_reports,
                "worksheet_count": worksheet_count,
                "processed_sheet_count": (
                    processed_sheet_count
                ),
                "skipped_sheet_count": (
                    skipped_sheet_count
                ),
                "worksheet_item_total": (
                    worksheet_item_total
                ),
                "total_items": len(all_items),
                "duplicate_count": (
                    workbook_duplicates
                ),
                "conflict_count": len(conflicts),
                "invalid_row_count": (
                    invalid_row_count
                ),
                "formula_price_row_count": (
                    formula_price_row_count
                ),
                "importer_version": (
                    EXCEL_IMPORTER_VERSION
                ),
                "rapidfuzz_available": (
                    RAPIDFUZZ_AVAILABLE
                ),
            }

        finally:
            workbook.close()
            formula_workbook.close()

            print(
                "[EXCEL] Workbooks closed.",
                flush=True,
            )
